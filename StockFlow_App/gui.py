"""StockFlow ERP Lite arayüzü."""
from __future__ import annotations

from collections import Counter
from datetime import datetime
from pathlib import Path

from PyQt6.QtCore import Qt
from PyQt6.QtGui import QAction, QDoubleValidator, QIntValidator
from PyQt6.QtWidgets import (
    QApplication, QFileDialog, QGroupBox, QHBoxLayout, QInputDialog,
    QLabel, QLineEdit, QMainWindow, QMenu, QMessageBox, QPushButton,
    QGridLayout, QTableWidget, QTableWidgetItem, QTabWidget, QVBoxLayout,
    QWidget, QHeaderView,
)

from database import (
    create_database, delete_product, get_dashboard_stats, get_product_by_barcode,
    get_products, get_sales, increase_stock, record_sale, save_product,
    update_product,
)
from models import Product

PRODUCT_HEADERS = ["Barkod", "Parça Adı", "Kategori", "Marka", "Araç", "OEM", "Raf", "Tedarikçi", "Alış", "Satış", "Stok", "Min."]
EXCEL_HEADERS = ["Barkod", "Parça Adı", "Kategori", "Marka", "Araç Modeli", "OEM Kodu", "Raf Kodu", "Tedarikçi", "Alış Fiyatı", "Satış Fiyatı", "Stok", "Min. Stok"]


class StockApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("StockFlow ERP Lite — Oto Yedek Parça Otomasyonu")
        self.resize(1420, 860)
        self.cart: list[Product] = []
        self._all_products: list[Product] = []
        create_database()
        self._init_ui()
        self.refresh_all()

    def _init_ui(self) -> None:
        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout(central)
        title = QLabel("StockFlow ERP Lite")
        title.setStyleSheet("font-size: 22px; font-weight: 700; color: #1e293b;")
        layout.addWidget(title)
        layout.addWidget(self._create_dashboard())

        self.tabs = QTabWidget()
        self.tabs.addTab(self._build_inventory_tab(), "Stok ve Satış")
        self.tabs.addTab(self._build_sales_tab(), "Geçmiş Satışlar")
        self.tabs.currentChanged.connect(lambda _: self._load_sales())
        layout.addWidget(self.tabs)

    def _create_dashboard(self) -> QWidget:
        widget = QWidget()
        row = QHBoxLayout(widget)
        self.lbl_total_types = self._stat_card("Toplam Çeşit", "#3b82f6")
        self.lbl_total_stock = self._stat_card("Toplam Stok", "#10b981")
        self.lbl_total_value = self._stat_card("Envanter Değeri", "#8b5cf6")
        self.lbl_low_stock = self._stat_card("Kritik Stok", "#ef4444")
        for card in (self.lbl_total_types, self.lbl_total_stock, self.lbl_total_value, self.lbl_low_stock):
            row.addWidget(card)
        return widget

    @staticmethod
    def _stat_card(title: str, color: str) -> QGroupBox:
        card = QGroupBox(title)
        card.setStyleSheet(f"QGroupBox {{font-weight:bold; border:2px solid {color}; border-radius:7px; margin-top:6px;}}")
        layout = QVBoxLayout(card)
        value = QLabel("0")
        value.setAlignment(Qt.AlignmentFlag.AlignCenter)
        value.setStyleSheet(f"font-size:18px; font-weight:bold; color:{color};")
        layout.addWidget(value)
        card.value_label = value
        return card

    def _build_inventory_tab(self) -> QWidget:
        tab = QWidget()
        layout = QVBoxLayout(tab)
        search = QHBoxLayout()
        self.txt_search = QLineEdit()
        self.txt_search.setPlaceholderText("Barkod okutun ve Enter ile sepete ekleyin")
        self.txt_search.returnPressed.connect(self._search_and_add_to_cart)
        search.addWidget(QLabel("Hızlı Satış / Barkod:"))
        search.addWidget(self.txt_search)
        layout.addLayout(search)

        middle = QHBoxLayout()
        middle.addWidget(self._build_product_form(), 3)
        middle.addWidget(self._build_cart(), 2)
        layout.addLayout(middle)

        actions = QHBoxLayout()
        import_button = QPushButton("Excel'den Toplu Yükle")
        import_button.clicked.connect(self._import_excel)
        export_button = QPushButton("Stoğu Excel'e Aktar")
        export_button.clicked.connect(self._export_excel)
        actions.addWidget(import_button)
        actions.addWidget(export_button)
        actions.addStretch()
        layout.addLayout(actions)

        group = QGroupBox("Stoktaki Yedek Parçalar")
        group_layout = QVBoxLayout(group)
        self.table = QTableWidget(0, len(PRODUCT_HEADERS))
        self.table.setHorizontalHeaderLabels(PRODUCT_HEADERS)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self._show_context_menu)
        self.table.cellDoubleClicked.connect(lambda row, _: self._fill_form_from_row(row))
        group_layout.addWidget(self.table)
        layout.addWidget(group)
        return tab

    def _build_product_form(self) -> QGroupBox:
        group = QGroupBox("Parça / Ürün Kartı")
        form = QGridLayout(group)
        self.txt_barcode, self.txt_name, self.txt_category, self.txt_brand = (QLineEdit() for _ in range(4))
        self.txt_vehicle, self.txt_oem, self.txt_shelf, self.txt_supplier = (QLineEdit() for _ in range(4))
        self.txt_purchase, self.txt_sale = QLineEdit("0"), QLineEdit("0")
        self.txt_stock, self.txt_min_stock = QLineEdit("0"), QLineEdit("5")
        decimal = QDoubleValidator(0.0, 999999999.0, 2, self)
        integer = QIntValidator(0, 999999999, self)
        self.txt_purchase.setValidator(decimal); self.txt_sale.setValidator(decimal)
        self.txt_stock.setValidator(integer); self.txt_min_stock.setValidator(integer)
        fields = [("Barkod / OEM:", self.txt_barcode), ("Parça Adı:", self.txt_name),
                  ("Kategori:", self.txt_category), ("Marka:", self.txt_brand),
                  ("Uyumlu Araç:", self.txt_vehicle), ("OEM No:", self.txt_oem),
                  ("Raf Kodu:", self.txt_shelf), ("Tedarikçi:", self.txt_supplier),
                  ("Alış Fiyatı:", self.txt_purchase), ("Satış Fiyatı:", self.txt_sale),
                  ("Stok Adedi:", self.txt_stock), ("Min. Stok:", self.txt_min_stock)]
        for index, (label, field) in enumerate(fields):
            row, column = divmod(index, 2)
            form.addWidget(QLabel(label), row, column * 2)
            form.addWidget(field, row, column * 2 + 1)
        buttons = QHBoxLayout()
        for label, handler in (("Ürün Ekle", self._add_product), ("Güncelle", self._update_product),
                               ("Sil", self._delete_product), ("Temizle", self._clear_form)):
            button = QPushButton(label); button.clicked.connect(handler); buttons.addWidget(button)
        form.addLayout(buttons, 6, 0, 1, 4)
        return group

    def _build_cart(self) -> QGroupBox:
        group = QGroupBox("Hızlı Satış Kasası")
        layout = QVBoxLayout(group)
        self.cart_table = QTableWidget(0, 3)
        self.cart_table.setHorizontalHeaderLabels(["Parça Adı", "Adet", "Fiyat"])
        self.cart_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.cart_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        layout.addWidget(self.cart_table)
        self.lbl_cart_total = QLabel("Toplam Tutar: 0,00 ₺")
        self.lbl_cart_total.setStyleSheet("font-size:16px; font-weight:bold;")
        layout.addWidget(self.lbl_cart_total)
        buttons = QHBoxLayout()
        complete = QPushButton("Satışı Tamamla"); complete.clicked.connect(self._complete_sale)
        clear = QPushButton("Sepeti Boşalt"); clear.clicked.connect(self._clear_cart)
        buttons.addWidget(complete); buttons.addWidget(clear); layout.addLayout(buttons)
        return group

    def _build_sales_tab(self) -> QWidget:
        tab = QWidget()
        layout = QVBoxLayout(tab)
        self.sales_summary = QLabel("Toplam satış: 0,00 ₺")
        self.sales_summary.setStyleSheet("font-size:16px; font-weight:bold;")
        layout.addWidget(self.sales_summary)
        self.sales_table = QTableWidget(0, 7)
        self.sales_table.setHorizontalHeaderLabels(["Fiş #", "Tarih", "Barkod", "Parça Adı", "Adet", "Birim Fiyat", "Toplam"])
        self.sales_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.sales_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.sales_table)
        refresh = QPushButton("Geçmişi Yenile"); refresh.clicked.connect(self._load_sales)
        layout.addWidget(refresh, alignment=Qt.AlignmentFlag.AlignRight)
        return tab

    def _product_from_form(self) -> Product:
        return Product(
            barcode=self.txt_barcode.text().strip(), product_name=self.txt_name.text().strip(),
            category=self.txt_category.text().strip(), brand=self.txt_brand.text().strip(),
            vehicle_model=self.txt_vehicle.text().strip(), oem_code=self.txt_oem.text().strip(),
            shelf_code=self.txt_shelf.text().strip(), supplier=self.txt_supplier.text().strip(),
            purchase_price=float(self.txt_purchase.text().replace(",", ".") or 0),
            sale_price=float(self.txt_sale.text().replace(",", ".") or 0),
            stock=int(self.txt_stock.text() or 0), min_stock=int(self.txt_min_stock.text() or 5),
        )

    def _add_product(self) -> None:
        try:
            product = self._product_from_form()
            if not product.barcode or not product.product_name:
                raise ValueError("Barkod ve parça adı zorunludur.")
            if not save_product(product):
                raise ValueError("Bu barkod zaten kayıtlı. Değişiklik için Güncelle'yi kullanın.")
            self.refresh_all(); self._clear_form()
            QMessageBox.information(self, "Başarılı", "Ürün eklendi.")
        except ValueError as error:
            QMessageBox.warning(self, "Kontrol edin", str(error))

    def _update_product(self) -> None:
        try:
            product = self._product_from_form()
            if not product.barcode or not product.product_name:
                raise ValueError("Barkod ve parça adı zorunludur.")
            if not update_product(product):
                raise ValueError("Güncellenecek barkod bulunamadı.")
            self.refresh_all(); QMessageBox.information(self, "Başarılı", "Ürün güncellendi.")
        except ValueError as error:
            QMessageBox.warning(self, "Kontrol edin", str(error))

    def _delete_product(self) -> None:
        barcode = self.txt_barcode.text().strip()
        if not barcode:
            QMessageBox.warning(self, "Uyarı", "Silmek için bir ürün seçin."); return
        if QMessageBox.question(self, "Ürünü sil", f"{barcode} barkodlu ürün silinsin mi?") != QMessageBox.StandardButton.Yes:
            return
        delete_product(barcode); self.refresh_all(); self._clear_form()

    def _clear_form(self) -> None:
        for field in (self.txt_barcode, self.txt_name, self.txt_category, self.txt_brand, self.txt_vehicle,
                      self.txt_oem, self.txt_shelf, self.txt_supplier): field.clear()
        self.txt_purchase.setText("0"); self.txt_sale.setText("0"); self.txt_stock.setText("0"); self.txt_min_stock.setText("5")

    def _fill_form_from_row(self, row: int) -> None:
        fields = (self.txt_barcode, self.txt_name, self.txt_category, self.txt_brand, self.txt_vehicle,
                  self.txt_oem, self.txt_shelf, self.txt_supplier, self.txt_purchase, self.txt_sale,
                  self.txt_stock, self.txt_min_stock)
        for column, field in enumerate(fields): field.setText(self.table.item(row, column).text())

    def _show_context_menu(self, position) -> None:
        row = self.table.rowAt(position.y())
        if row < 0: return
        self.table.selectRow(row)
        menu = QMenu(self)
        copy_action = menu.addAction("Barkodu Kopyala")
        increase_action = menu.addAction("Stok Artır")
        edit_action = menu.addAction("Düzenle")
        delete_action = menu.addAction("Sil")
        chosen = menu.exec(self.table.viewport().mapToGlobal(position))
        barcode = self.table.item(row, 0).text()
        if chosen == copy_action:
            QApplication.clipboard().setText(barcode)
        elif chosen == increase_action:
            amount, ok = QInputDialog.getInt(self, "Stok Artır", "Eklenecek adet:", 1, 1)
            if ok: increase_stock(barcode, amount); self.refresh_all()
        elif chosen == edit_action:
            self._fill_form_from_row(row)
        elif chosen == delete_action:
            self._fill_form_from_row(row); self._delete_product()

    def _load_products(self) -> None:
        self._all_products = get_products()
        self.table.setRowCount(len(self._all_products))
        for row, product in enumerate(self._all_products):
            values = (product.barcode, product.product_name, product.category, product.brand, product.vehicle_model,
                      product.oem_code, product.shelf_code, product.supplier, f"{product.purchase_price:.2f}",
                      f"{product.sale_price:.2f}", str(product.stock), str(product.min_stock))
            for column, value in enumerate(values): self.table.setItem(row, column, QTableWidgetItem(value))

    def _update_dashboard(self) -> None:
        stats = get_dashboard_stats()
        self.lbl_total_types.value_label.setText(str(stats["total_products"]))
        self.lbl_total_stock.value_label.setText(str(stats["total_stock"]))
        self.lbl_total_value.value_label.setText(f"{stats['total_value']:,.2f} ₺")
        self.lbl_low_stock.value_label.setText(str(stats["low_stock_count"]))

    def refresh_all(self) -> None:
        self._load_products(); self._update_dashboard(); self._load_sales()

    def _search_and_add_to_cart(self) -> None:
        barcode = self.txt_search.text().strip()
        if not barcode: return
        product = get_product_by_barcode(barcode)
        if product is None:
            QMessageBox.warning(self, "Bulunamadı", "Bu barkoda ait ürün bulunamadı."); return
        self.cart.append(product); self.txt_search.clear(); self._update_cart_table()

    def _update_cart_table(self) -> None:
        counts = Counter(product.barcode for product in self.cart)
        products = {product.barcode: product for product in self.cart}
        self.cart_table.setRowCount(len(counts))
        total = 0.0
        for row, (barcode, quantity) in enumerate(counts.items()):
            product = products[barcode]; line_total = product.sale_price * quantity; total += line_total
            for column, value in enumerate((product.product_name, str(quantity), f"{line_total:.2f} ₺")):
                self.cart_table.setItem(row, column, QTableWidgetItem(value))
        self.lbl_cart_total.setText(f"Toplam Tutar: {total:.2f} ₺")

    def _clear_cart(self) -> None:
        self.cart.clear(); self._update_cart_table()

    def _complete_sale(self) -> None:
        if not self.cart:
            QMessageBox.warning(self, "Sepet Boş", "Sepette ürün yok."); return
        counts = Counter(product.barcode for product in self.cart)
        products = {product.barcode: product for product in self.cart}
        items = [(products[barcode], quantity) for barcode, quantity in counts.items()]
        try:
            record_sale(items)
        except ValueError as error:
            QMessageBox.warning(self, "Satış yapılamadı", str(error)); return
        total = sum(product.sale_price * quantity for product, quantity in items)
        receipt = "=== StockFlow ERP Lite Fişi ===\n" + f"Tarih: {datetime.now():%d.%m.%Y %H:%M}\n" + "-" * 30 + "\n"
        receipt += "\n".join(f"{product.product_name} x{quantity} — {product.sale_price * quantity:.2f} ₺" for product, quantity in items)
        receipt += f"\n{'-' * 30}\nTOPLAM: {total:.2f} ₺"
        self._clear_cart(); self.refresh_all()
        QMessageBox.information(self, "Satış Tamamlandı", receipt)

    def _load_sales(self) -> None:
        sales = get_sales()
        self.sales_table.setRowCount(len(sales))
        total = 0.0
        for row, sale in enumerate(sales):
            total += sale["total"]
            values = (str(sale["id"]), sale["sale_date"], sale["barcode"], sale["product_name"],
                      str(sale["quantity"]), f"{sale['sale_price']:.2f} ₺", f"{sale['total']:.2f} ₺")
            for column, value in enumerate(values): self.sales_table.setItem(row, column, QTableWidgetItem(value))
        self.sales_summary.setText(f"Toplam satış: {total:.2f} ₺ • {len(sales)} satır")

    def _export_excel(self) -> None:
        path, _ = QFileDialog.getSaveFileName(self, "Stoğu Excel'e Aktar", "stockflow_stok.xlsx", "Excel Dosyası (*.xlsx)")
        if not path: return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            workbook = Workbook(); sheet = workbook.active; sheet.title = "Stok"
            sheet.append(EXCEL_HEADERS)
            for product in get_products():
                sheet.append([product.barcode, product.product_name, product.category, product.brand, product.vehicle_model,
                              product.oem_code, product.shelf_code, product.supplier, product.purchase_price,
                              product.sale_price, product.stock, product.min_stock])
            for cell in sheet[1]: cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor="1E3A5F")
            sheet.freeze_panes = "A2"; sheet.auto_filter.ref = sheet.dimensions
            for column in sheet.columns:
                sheet.column_dimensions[column[0].column_letter].width = min(max(len(str(cell.value or "")) for cell in column) + 2, 30)
            workbook.save(path)
            QMessageBox.information(self, "Başarılı", f"{len(self._all_products)} ürün Excel'e aktarıldı.")
        except Exception as error:
            QMessageBox.critical(self, "Excel hatası", f"Dosya oluşturulamadı:\n{error}")

    def _import_excel(self) -> None:
        path, _ = QFileDialog.getOpenFileName(self, "Excel'den Toplu Ürün Yükle", "", "Excel Dosyası (*.xlsx)")
        if not path: return
        try:
            from openpyxl import load_workbook
            sheet = load_workbook(path, data_only=True).active
            headers = [str(cell.value or "").strip().casefold() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            expected = [header.casefold() for header in EXCEL_HEADERS]
            if headers[:len(expected)] != expected:
                raise ValueError("Başlıklar şu sırada olmalı: " + ", ".join(EXCEL_HEADERS))
            added = updated = skipped = 0
            for values in sheet.iter_rows(min_row=2, values_only=True):
                if not any(value is not None and str(value).strip() for value in values): continue
                try:
                    value = list(values) + [None] * 12
                    product = Product(
                        barcode=str(value[0]).strip(), product_name=str(value[1]).strip(),
                        category=str(value[2] or "").strip(), brand=str(value[3] or "").strip(),
                        vehicle_model=str(value[4] or "").strip(), oem_code=str(value[5] or "").strip(),
                        shelf_code=str(value[6] or "").strip(), supplier=str(value[7] or "").strip(),
                        purchase_price=float(value[8] or 0), sale_price=float(value[9] or 0),
                        stock=int(value[10] or 0), min_stock=int(value[11] or 5),
                    )
                    if not product.barcode or not product.product_name: raise ValueError
                    if save_product(product): added += 1
                    else: update_product(product); updated += 1
                except (ValueError, TypeError): skipped += 1
            self.refresh_all()
            QMessageBox.information(self, "İçe aktarma tamamlandı", f"Eklenen: {added}\nGüncellenen: {updated}\nAtlanan: {skipped}")
        except Exception as error:
            QMessageBox.critical(self, "Excel hatası", f"Dosya okunamadı:\n{error}")

